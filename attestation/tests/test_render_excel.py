from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from attestation.services.render import render_excel_5


class RenderExcel5Tests(unittest.TestCase):
    def test_subdivision_header_is_emitted_before_each_group(self):
        def workplace(no: str, subdivision: str) -> dict:
            return {
                "workplace_no": no,
                "position": f"Position {no}",
                "subdivision": subdivision,
                "em_measurements": {
                    "1.4.10": {
                        "norma": "2.5",
                        "actual": "0.36",
                        "time": "20",
                        "cls": "2.0",
                    },
                },
                "factors": {"em_field": "2.0"},
            }

        workplaces = [
            workplace("000001", "First department"),
            workplace("000002", "First department"),
            workplace("000003", "Second department"),
        ]

        with tempfile.TemporaryDirectory() as tmp:
            output = render_excel_5({}, workplaces, Path(tmp) / "excel_5.xlsx", lang="lat")
            ws = load_workbook(output, data_only=False)["complete"]

        self.assertEqual(ws["A27"].value, "First department")
        self.assertEqual(ws["A28"].value, "000001")
        self.assertEqual(ws["A44"].value, "000002")
        self.assertEqual(ws["A60"].value, "Second department")
        self.assertEqual(ws["A61"].value, "000003")
        merges = {str(cell_range) for cell_range in ws.merged_cells.ranges}
        self.assertIn("A27:K27", merges)
        self.assertIn("A60:K60", merges)


if __name__ == "__main__":
    unittest.main()
