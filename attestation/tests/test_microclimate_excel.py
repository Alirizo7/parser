from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from attestation.services import xlsx
from attestation.services.extract import (
    FactorRow,
    _extract_microclimate_measurements,
    _parse_factor_rows,
)
from attestation.services.render import _subrows_file3, render_excel_3


def _factor_row(
    section: str,
    *,
    name: str = "",
    norma: str = "",
    actual: str = "-",
    duration: str = "",
    cls: str = "",
) -> FactorRow:
    """Build the relevant part of a factor-table row for microclimate tests."""
    return FactorRow(
        section=section,
        c0=section,
        name=name,
        norm=norma,
        actual=actual,
        duration=duration,
        cls=cls,
    )


class MicroclimateExtractionTests(unittest.TestCase):
    def test_warm_period_1_8_keeps_existing_measurement_mapping(self):
        rows = [
            _factor_row("1.7.2", name="Iб – 88(78-97), Vt/m"),
            _factor_row(
                "1.8.2", name="Ҳарорат", norma="20,0-24,0", actual="21",
                duration="80", cls="2",
            ),
            _factor_row(
                "1.8.6", name="Иссиқлик тарқалиши", norma="100", actual="йўқ",
            ),
            _factor_row(
                "1.8.7", name="Ҳаво ҳаракати тезлиги", norma="<0,2", actual="0,1",
                duration="80", cls="2",
            ),
            _factor_row(
                "1.8.8", name="Ҳавонинг нисбий намлиги", norma="75", actual="56",
                duration="80", cls="2",
            ),
        ]

        result = _extract_microclimate_measurements(rows)

        self.assertEqual(result["category_label"], "Iб – 88(78-97)")
        self.assertEqual(result["temp"], {
            "norma": "20,0-24,0", "actual": "21", "time": "80", "cls": "2.0",
        })
        self.assertEqual(result["air_speed"], {
            "norma": "<0,2", "actual": "0,1", "time": "80", "cls": "2.0",
        })
        self.assertEqual(result["humidity"], {
            "norma": "75", "actual": "56", "time": "80", "cls": "2.0",
        })
        self.assertEqual(result["heat_radiation"], {
            "norma": "100", "actual": "йўқ", "time": "", "cls": "-",
        })

    def test_cold_indoor_1_9_maps_all_measurements_and_category(self):
        rows = [
            # The supplied cold-period cards retain the category captions in 1.7.
            _factor_row("1.7.4", name="IIб – 233(175-290), Vt/m"),
            # Warm-period branch is present in the card, but intentionally empty.
            _factor_row("1.8.4", name="Ҳарорат", norma="23,0-26,0", actual="-"),
            _factor_row(
                "1.9.4", name="Ҳарорат", norma="17,0-19,0", actual="18",
                duration="80", cls="2",
            ),
            _factor_row(
                "1.9.6", name="Ҳаво ҳаракати тезлиги", norma="<0,4", actual="0,3",
                duration="80", cls="2",
            ),
            _factor_row(
                "1.9.7", name="Ҳавонинг нисбий намлиги", norma="75", actual="53",
                duration="80", cls="2",
            ),
            _factor_row(
                "1.9.8", name="Иссиқлик тарқалиши", norma="100", actual="йўқ",
            ),
        ]

        result = _extract_microclimate_measurements(rows)

        self.assertEqual(result["category_label"], "IIб – 233(175-290)")
        self.assertEqual(result["temp"], {
            "norma": "17,0-19,0", "actual": "18", "time": "80", "cls": "2.0",
        })
        self.assertEqual(result["air_speed"], {
            "norma": "<0,4", "actual": "0,3", "time": "80", "cls": "2.0",
        })
        self.assertEqual(result["humidity"], {
            "norma": "75", "actual": "53", "time": "80", "cls": "2.0",
        })
        self.assertEqual(result["heat_radiation"], {
            "norma": "100", "actual": "йўқ", "time": "", "cls": "-",
        })

    def test_cold_outdoor_1_10_1_is_emitted_in_seventh_excel_subrow(self):
        measurements = _extract_microclimate_measurements([
            _factor_row(
                "1.10.1", name="Ҳарорат (очиқ ҳудуд)", norma="-4,2", actual="5",
                duration="80", cls="2",
            ),
        ])

        subrows = _subrows_file3({"microclimate_measurements": measurements})

        self.assertEqual(len(subrows), 9)
        outdoor_row = subrows[6]
        self.assertEqual(outdoor_row[4], -4.2)  # D: hygienic norm
        self.assertIs(outdoor_row[5], xlsx.KEEP)
        self.assertIs(outdoor_row[6], xlsx.KEEP)
        self.assertIs(outdoor_row[7], xlsx.KEEP)
        self.assertEqual(outdoor_row[8], 5)     # H: actual value
        self.assertEqual(outdoor_row[9], 80)    # I: exposure time
        self.assertEqual(outdoor_row[10], 2)    # J: class

        workplace = {
            "workplace_no": "000005",
            "position": "Outdoor worker",
            "subdivision": "Road section",
            "microclimate_measurements": measurements,
            "factors": {"microclimate": "2"},
        }
        with tempfile.TemporaryDirectory() as tmp:
            output = render_excel_3(
                {}, [workplace], Path(tmp) / "excel_3.xlsx", lang="cyr",
            )
            ws = load_workbook(output, data_only=False)["complete"]

        # Row 28 starts the only 9-row block; outdoor temperature is offset 6.
        self.assertEqual(ws["D34"].value, -4.2)
        self.assertEqual(ws["E34"].value, "=H34+0.1")
        self.assertEqual(ws["F34"].value, "=H34+0.2")
        self.assertEqual(ws["G34"].value, "=H34-0.3")
        self.assertEqual(ws["H34"].value, 5)
        self.assertEqual(ws["I34"].value, 80)
        self.assertEqual(ws["J34"].value, 2)

    def test_missing_microclimate_data_keeps_all_measurement_cells_empty(self):
        measurements = _extract_microclimate_measurements([
            _factor_row("1.8.1"),
            _factor_row("1.9.1"),
            _factor_row("1.10.1"),
        ])

        subrows = _subrows_file3({"microclimate_measurements": measurements})

        self.assertEqual(measurements["category_label"], "")
        self.assertIsNone(measurements["temp"])
        self.assertIsNone(measurements["air_speed"])
        self.assertIsNone(measurements["humidity"])
        self.assertIsNone(measurements["heat_radiation"])
        self.assertIsNone(measurements["outdoor_temp"])
        self.assertTrue(all(
            subrow[column] is xlsx.CLEAR
            for subrow in subrows
            for column in range(4, 11)
        ))

    def test_outdoor_subtotal_inheriting_1_10_1_is_not_a_measurement(self):
        measurements = _extract_microclimate_measurements([
            _factor_row(
                "1.10.1", name="Ҳаво ҳарорати, °С", norma="", actual="-",
            ),
            # In converted cards the section number is inherited from the row
            # above, while merged cells repeat the subtotal caption in every
            # parsed column, including ``actual``.
            _factor_row(
                "1.10.1", name="Микроиқлим умумий баҳолаш",
                norma="Микроиқлим умумий баҳолаш",
                actual="Микроиқлим умумий баҳолаш", cls="2.0",
            ),
        ])

        self.assertIsNone(measurements["outdoor_temp"])

    def test_numeric_heat_radiation_gets_numeric_pseudo_measurements(self):
        measurements = _extract_microclimate_measurements([
            _factor_row("1.7.5", name="III – 177(161-193), Vt/m"),
            _factor_row(
                "1.9.5", name="Ҳаво ҳарорати", norma="13,0-19,0",
                actual="18", duration="80", cls="2",
            ),
            _factor_row(
                "1.9.8", name="Иссиқлик тарқалиши", norma="140",
                actual="135", duration="80", cls="2",
            ),
        ])
        workplace = {
            "workplace_no": "000013",
            "position": "Welder",
            "subdivision": "Plant",
            "microclimate_measurements": measurements,
            "factors": {"microclimate": "2"},
        }

        with tempfile.TemporaryDirectory() as tmp:
            output = render_excel_3(
                {}, [workplace], Path(tmp) / "excel_3.xlsx", lang="cyr",
            )
            ws = load_workbook(output, data_only=False)["complete"]

        # Heat radiation is offset 5 from the first block row 28.
        self.assertEqual(ws["D33"].value, 140)
        self.assertEqual(ws["E33"].value, "=H33+1")
        self.assertEqual(ws["F33"].value, "=H33+2")
        self.assertEqual(ws["G33"].value, "=H33-3")
        self.assertEqual(ws["H33"].value, 135)
        self.assertEqual(ws["I33"].value, 80)
        self.assertEqual(ws["J33"].value, 2)


class FactorColumnDetectionTests(unittest.TestCase):
    def test_hyphenated_duration_header_does_not_reuse_actual_column(self):
        grid = [
            [
                "Т/р", "Меҳнат жараёни ва ишлаб чиқариш муҳитининг омиллари",
                "Гигиеник меъёр", "Гигиеник меъёр", "Ҳақиқий даражаси",
                "Таъсир этиш давомий-лиги", "Таъсир этиш давомий-лиги",
                "Меҳнат шароитлари класси",
            ],
            [
                "1.9.4", "Ҳаво ҳарорати", "Нормада эмас", "Нормада эмас",
                "23,0", "", "", "",
            ],
        ]

        row = _parse_factor_rows(grid)[1]

        self.assertEqual(row.actual, "23,0")
        self.assertEqual(row.duration, "")
        self.assertNotEqual(row.duration, row.actual)

    def test_hyphenated_merged_duration_uses_last_duplicate_column(self):
        grid = [
            [
                "Т/р", "Меҳнат жараёни ва ишлаб чиқариш муҳитининг омиллари",
                "Гигиеник меъёр", "Ҳақиқий даражаси",
                "Таъсир этиш давомий-лиги", "Таъсир этиш давомий-лиги",
                "Меҳнат шароитлари класси",
            ],
            ["1.9.2", "Ҳаво ҳарорати", "20,0-24,0", "21", "80", "80", "2"],
        ]

        row = _parse_factor_rows(grid)[1]

        self.assertEqual(row.norm, "20,0-24,0")
        self.assertEqual(row.actual, "21")
        self.assertEqual(row.duration, "80")
        self.assertEqual(row.cls, "2")


if __name__ == "__main__":
    unittest.main()
