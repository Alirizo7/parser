"""Низкоуровневые помощники генерации xlsx-протоколов через openpyxl.

Подход тот же, что у docx-рендеров: берём пустой шаблон-ассет (шапка + заголовки
колонок + ОДИН прото-заголовок группы + ОДИН прото-блок рабочего места), снимаем
прототипы, вычищаем тело и генерируем его заново — клонируя прото-строки с полным
форматированием (стили ячеек, границы, объединения, высоты строк) и подставляя
значения. Псевдозамеры (колонки «1/2/3-фаолият») — ФОРМУЛЫ из прото-строки; при
переносе на новую строку относительные ссылки правит ``Translator`` (шаг ±0.1 и
т.п. НЕ хардкодим — он зашит в формуле прото-строки).

Транслитерация вывода — как в docx: держим управляемую часть листа (заголовки
таблицы + тело) на КИРИЛЛИЦЕ, при ``lang == 'lat'`` переводим её ``to_latin``.
Статичная шапка (приборы/НД/лаборатория) остаётся «как есть» (двуязычный
boilerplate клиента), меняется только заказчик.
"""
from __future__ import annotations

import re
from copy import copy

from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter

from .normalize import to_cyrillic, to_latin

# Ячейки, которые НЕ переводим в кириллицу: римские числа (разряд зрит. работ
# «III»/«IV»/«VI») и метка категории микроклимата («Ib – 88(78-97)»). ``to_cyrillic``
# исказил бы их (I→и, V→в → «ИВ»), а ``to_latin`` их не трогает (они уже латиница).
_KEEP_LATIN_RE = re.compile(r"^[IVXLCDM]+[а-яёa-z]?(\s*[–—-].*)?$")

KEEP = object()   # оставить значение прото-ячейки (формулу — с правкой ссылок)
CLEAR = object()  # очистить ячейку (пустая рамка)


def col(c: int) -> str:
    return get_column_letter(c)


# --- Снятие прото-строк -----------------------------------------------------
def capture_row(ws, row: int, ncols: int) -> dict:
    """Снять прототип строки: значение/формула + стиль каждой ячейки + высота."""
    cells = []
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cells.append({"value": cell.value, "style": copy(cell._style),
                      "coord": cell.coordinate})
    return {"cells": cells, "height": ws.row_dimensions[row].height}


def capture_merges(ws, r0: int, r1: int) -> list[tuple[int, int, int, int]]:
    """Объединения, целиком лежащие в строках [r0, r1], как относительные смещения."""
    out = []
    for mr in ws.merged_cells.ranges:
        if mr.min_row >= r0 and mr.max_row <= r1:
            out.append((mr.min_row - r0, mr.min_col, mr.max_row - r0, mr.max_col))
    return out


# --- Запись строк по прототипу ----------------------------------------------
def emit_row(ws, dst_row: int, proto: dict, ncols: int, values: dict | None = None) -> None:
    """Записать строку в ``dst_row`` по прототипу.

    ``values`` — переопределения по номеру колонки (1-индекс): конкретное
    значение, ``KEEP`` (взять из прото, формулу перенести Translator-ом) или
    ``CLEAR`` (очистить). Колонки без записи по умолчанию ведут себя как ``KEEP``.
    """
    values = values or {}
    for c in range(1, ncols + 1):
        cell = ws.cell(row=dst_row, column=c)
        pc = proto["cells"][c - 1]
        cell._style = copy(pc["style"])
        override = values.get(c, KEEP)
        if override is CLEAR:
            cell.value = None
        elif override is KEEP:
            cell.value = _kept_value(pc, c, dst_row)
        else:
            cell.value = override
    if proto["height"] is not None:
        ws.row_dimensions[dst_row].height = proto["height"]


def _kept_value(pc: dict, c: int, dst_row: int):
    """Значение прото-ячейки; формулу переносим на строку ``dst_row``."""
    v = pc["value"]
    if isinstance(v, str) and v.startswith("="):
        return Translator(v, origin=pc["coord"]).translate_formula(f"{col(c)}{dst_row}")
    return v


def apply_merges(ws, merges, dst_r0: int) -> None:
    for dr0, c0, dr1, c1 in merges:
        ws.merge_cells(start_row=dst_r0 + dr0, start_column=c0,
                       end_row=dst_r0 + dr1, end_column=c1)


def merge_span(ws, r0: int, r1: int, c: int) -> None:
    """Объединить одну колонку ``c`` по строкам [r0, r1] (если больше одной)."""
    if r1 > r0:
        ws.merge_cells(start_row=r0, start_column=c, end_row=r1, end_column=c)


def clear_body(ws, from_row: int) -> None:
    """Убрать объединения и строки тела начиная с ``from_row`` (ниже прото-блока).

    Снимаем ЛЮБОЕ объединение, задевающее удаляемую область (``max_row >=
    from_row``), в т.ч. вертикальный спан блока, начатый выше границы (иначе
    openpyxl ``delete_rows`` оставит «висящее» объединение).
    """
    for mr in list(ws.merged_cells.ranges):
        if mr.max_row >= from_row:
            ws.unmerge_cells(str(mr))
    if ws.max_row >= from_row:
        ws.delete_rows(from_row, ws.max_row - from_row + 1)


# --- Число / текст ----------------------------------------------------------
def to_number(value):
    """«0,7»→0.7, «27,0»→27.0, «50»→50; диапазон/текст/«-»→строка; пусто→None.

    Нужно, чтобы факт (H) был ЧИСЛОМ — иначе формулы псевдозамеров «=H+0.1» не
    посчитаются. Диапазоны норм («23,0-31,0») и текст («йўқ») остаются строками.
    """
    if value is None:
        return None
    s = str(value).strip()
    if s == "" or s == "-":
        return None
    norm = s.replace(",", ".")
    try:
        f = float(norm)
    except ValueError:
        return s  # диапазон / текст
    return int(f) if f.is_integer() else f


# --- Транслитерация управляемой части листа ---------------------------------
def transliterate_region(ws, lang: str, start_row: int) -> None:
    """Привести текст управляемой части листа (со ``start_row``) к письму ``lang``.

    Транслитерация «к целевому письму»: ``to_latin``/``to_cyrillic`` НЕ трогают
    текст, уже написанный в целевом письме (``to_latin`` — no-op для латиницы,
    ``to_cyrillic`` — для кириллицы), поэтому подход работает независимо от того,
    на каком письме исходно свёрстан ассет, и для латинских, и для кириллических
    карт. Ассет-эталон здесь на ЛАТИНИЦЕ → при ``lang='lat'`` метки остаются точь-
    в-точь как в эталоне (никакого round-trip «energiya→yenergiya»), при
    ``lang='cyr'`` переводятся в кириллицу. Статичную шапку (строки выше
    ``start_row``) не трогаем — это двуязычный boilerplate «как есть».
    """
    conv = {"lat": to_latin, "cyr": to_cyrillic}.get(lang)
    if conv is None:
        return
    protect = conv is to_cyrillic  # беречь римские/категорию только при cyr
    for row in ws.iter_rows(min_row=start_row):
        for cell in row:
            v = cell.value
            if isinstance(v, str) and not v.startswith("="):
                if protect and _KEEP_LATIN_RE.match(v.strip()):
                    continue
                cell.value = conv(v)
