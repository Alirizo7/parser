"""Построить пустые шаблоны-ассеты Excel-протоколов из эталонов клиента.

Из каждого ``excel_templates/N_*.xlsx`` (эталон, заполненный данными «Суд-
экспертиза Қарши») делаем ``attestation/assets/template_excel_N.xlsx``:
    * оставляем шапку (строки 1..~25) + заголовки колонок + ОДИН прото-заголовок
      группы (файлы 1–4) + ОДИН прото-блок рабочего места;
    * удаляем данные всех остальных РМ; удаляем лишний лист «Лист1» (файлы 4/5);
    * переводим управляемую часть листа (заголовки таблицы + прото-блок) в
      кириллицу — прото-шаблон держим на ОДНОМ письме, рендер транслитерирует
      по output_lang (как docx-рендеры).

Запуск:  python scripts/build_excel_assets.py
"""
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from attestation.services.xlsx import clear_body

# managed_start — первая строка заголовков таблицы (ниже — управляемая часть);
# group_row — прото-заголовок подразделения (None у файла 5);
# block_start / proto_len — прото-блок РМ (файл 1: 1 прото-строка вещества).
CFG = {
    1: dict(src="1_vrednie_veshestva",   managed_start=27, group_row=30, block_start=31, proto_len=1),
    2: dict(src="2_fizicheskie",         managed_start=27, group_row=30, block_start=31, proto_len=4),
    3: dict(src="3_mikroklimat",         managed_start=24, group_row=27, block_start=28, proto_len=9),
    4: dict(src="4_osveshennost",        managed_start=24, group_row=27, block_start=28, proto_len=5),
    5: dict(src="5_elektromagnitnie",    managed_start=24, group_row=None, block_start=28, proto_len=16),
}

SRC_DIR = ROOT / "excel_templates"
OUT_DIR = ROOT / "attestation" / "assets"


def build(idx: int, cfg: dict) -> None:
    wb = load_workbook(SRC_DIR / f"{cfg['src']}.xlsx")
    # Удаляем мусорный лист чужого клиента (файлы 4/5)
    for name in list(wb.sheetnames):
        if name != "complete":
            del wb[name]
    ws = wb["complete"]

    keep_until = cfg["block_start"] + cfg["proto_len"] - 1
    clear_body(ws, keep_until + 1)          # оставить header+colhdr+group+прото-блок
    # Управляемую часть НЕ переводим: держим ассет на письме эталона (латиница).
    # Рендер приводит её к output_lang «к целевому письму» (см. xlsx.transliterate_region).

    out = OUT_DIR / f"template_excel_{idx}.xlsx"
    wb.save(out)
    print(f"file{idx}: {out.name}  (proto rows {cfg['block_start']}..{keep_until}, "
          f"max_row now {ws.max_row})")


if __name__ == "__main__":
    for idx, cfg in CFG.items():
        build(idx, cfg)
